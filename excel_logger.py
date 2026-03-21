"""
Excel logger for options trading bot.

Tabs:
    Small         — open positions for small portfolio
    Medium        — open positions for medium portfolio
    Large         — open positions for large portfolio
    Trade History — all closed trades across all portfolios

Public functions:
    log_pipeline_run()   — called by main.py after each scan (scan log)
    update_positions()   — rewrites open position tabs from paper_portfolio.json
    log_closed_trade()   — appends a closed trade to Trade History tab
    init_workbook()      — creates trades.xlsx with correct tabs if it doesn't exist
"""

import json
from pathlib import Path
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

EXCEL_PATH     = Path("trades.xlsx")
PORTFOLIO_TABS = ["Small", "Medium", "Large"]
HISTORY_TAB    = "Trade History"

# Colors
COLOR_HEADER_BG  = "1F4E79"   # Dark blue
COLOR_HEADER_FG  = "FFFFFF"   # White
COLOR_PROFIT_BG  = "E2EFDA"   # Light green
COLOR_LOSS_BG    = "FCE4D6"   # Light red / orange
COLOR_NEUTRAL_BG = "FFFFFF"   # White
COLOR_ALT_ROW    = "F2F2F2"   # Light grey alternate row
COLOR_SUBHEADER  = "D6E4F0"   # Light blue for section labels

# Fonts
FONT_HEADER  = Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=10)
FONT_TITLE   = Font(name="Arial", bold=True, size=12)
FONT_NORMAL  = Font(name="Arial", size=10)
FONT_BOLD    = Font(name="Arial", bold=True, size=10)

# Thin border for cells
_THIN = Side(style="thin", color="CCCCCC")
BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Column definitions for open positions tabs
POSITION_COLUMNS = [
    ("Position ID",      12),
    ("Symbol",           8),
    ("Type",             6),
    ("Strategy",         8),
    ("Strike",           8),
    ("Expiration",       12),
    ("DTE",              5),
    ("Entry Date",       16),
    ("Entry Premium",    14),
    ("Current Premium",  15),
    ("Unrealized P&L",   14),
    ("P&L %",            8),
    ("Take Profit At",   14),
    ("Stop Loss At",     12),
    ("Capital Required", 16),
    ("Status",           10),
]

# Column definitions for trade history tab
HISTORY_COLUMNS = [
    ("Position ID",    12),
    ("Portfolio",      10),
    ("Symbol",          8),
    ("Type",            6),
    ("Strategy",        8),
    ("Strike",          8),
    ("Expiration",     12),
    ("Entry Date",     16),
    ("Exit Date",      16),
    ("Days Held",       9),
    ("Entry Premium",  14),
    ("Exit Premium",   13),
    ("Realized P&L",   13),
    ("P&L %",           8),
    ("Close Reason",   16),
    ("Win / Loss",     10),
]

# Column definitions for scan log (written by log_pipeline_run)
SCAN_COLUMNS = [
    ("Timestamp",        18),
    ("Portfolio",        10),
    ("Symbol",            8),
    ("Type",              6),
    ("Strike",            8),
    ("Bid",               8),
    ("Delta",             7),
    ("Spread",            8),
    ("Underlying Price", 17),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", start_color=color, fgColor=color)


def _row_fill(color: str) -> PatternFill:
    return PatternFill("solid", start_color=color, fgColor=color)


def _write_header_row(ws, columns: list, row: int = 1):
    """Write a styled header row from a list of (label, width) tuples."""
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font      = FONT_HEADER
        cell.fill      = _header_fill(COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[row].height = 28


def _write_title(ws, title: str, col_count: int):
    """Write a merged title row above the headers."""
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font      = FONT_TITLE
    cell.fill      = _header_fill(COLOR_SUBHEADER)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22


def _style_data_row(ws, row: int, col_count: int, pnl: Optional[float] = None):
    """Apply background color and border to a data row."""
    if pnl is not None and pnl > 0:
        bg = COLOR_PROFIT_BG
    elif pnl is not None and pnl < 0:
        bg = COLOR_LOSS_BG
    elif row % 2 == 0:
        bg = COLOR_ALT_ROW
    else:
        bg = COLOR_NEUTRAL_BG

    fill = _row_fill(bg)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = fill
        cell.border = BORDER_THIN
        cell.font   = FONT_NORMAL
        cell.alignment = Alignment(vertical="center")


def _currency(val) -> str:
    """Format a number as a currency string, or return '-' if None."""
    if val is None:
        return "-"
    try:
        return f"${float(val):,.2f}"
    except (TypeError, ValueError):
        return "-"


def _pct(val) -> str:
    """Format a number as a percentage string."""
    if val is None:
        return "-"
    try:
        return f"{float(val):+.1f}%"
    except (TypeError, ValueError):
        return "-"


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")


# ---------------------------------------------------------------------------
# Workbook initialisation
# ---------------------------------------------------------------------------

def init_workbook() -> Workbook:
    """
    Create trades.xlsx with all required tabs if it doesn't exist.
    If it already exists, load and return it.
    """
    if EXCEL_PATH.exists():
        return load_workbook(EXCEL_PATH)

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create portfolio tabs
    for tab_name in PORTFOLIO_TABS:
        ws = wb.create_sheet(tab_name)
        _write_header_row(ws, POSITION_COLUMNS, row=1)
        _write_title(ws, f"{tab_name} Portfolio — Open Positions", len(POSITION_COLUMNS))
        ws.freeze_panes = "A3"   # Freeze title + header rows

    # Create Trade History tab
    ws = wb.create_sheet(HISTORY_TAB)
    _write_header_row(ws, HISTORY_COLUMNS, row=1)
    _write_title(ws, "Trade History — All Portfolios", len(HISTORY_COLUMNS))
    ws.freeze_panes = "A3"

    # Create Scan Log tab
    ws = wb.create_sheet("Scan Log")
    _write_header_row(ws, SCAN_COLUMNS, row=1)
    _write_title(ws, "Scan Log — Eligible Options Found Each Run", len(SCAN_COLUMNS))
    ws.freeze_panes = "A3"

    wb.save(EXCEL_PATH)
    print(f"[ExcelLogger] Created {EXCEL_PATH}")
    return wb


# ---------------------------------------------------------------------------
# Update open positions tabs
# ---------------------------------------------------------------------------

def update_positions(state_file: str = "paper_portfolio.json"):
    """
    Rewrite the Small, Medium, Large tabs from current paper_portfolio.json state.
    Clears and rewrites each tab completely so data is always fresh.
    """
    if not Path(state_file).exists():
        print("[ExcelLogger] paper_portfolio.json not found — skipping position update.")
        return

    with open(state_file, "r") as f:
        state = json.load(f)

    wb = init_workbook()
    open_positions = [p for p in state.get("open_positions", []) if p["status"] == "OPEN"]

    for tab_name in PORTFOLIO_TABS:
        ws = wb[tab_name]

        # Clear existing data rows (keep title row 1 + header row 2)
        if ws.max_row > 2:
            ws.delete_rows(3, ws.max_row)

        # Filter positions for this portfolio
        portfolio_name = tab_name.lower()
        positions = [p for p in open_positions if p["portfolio_name"] == portfolio_name]

        # Update title with last updated timestamp
        ws.cell(row=1, column=1).value = (
            f"{tab_name} Portfolio — Open Positions   "
            f"(Last updated: {_timestamp()})   "
            f"Open: {len(positions)}"
        )

        for pos in positions:
            row = ws.max_row + 1
            pnl = pos.get("unrealized_pnl")

            values = [
                pos.get("position_id", ""),
                pos.get("symbol", ""),
                pos.get("option_type", ""),
                pos.get("strategy", ""),
                pos.get("strike", ""),
                pos.get("expiration_date", ""),
                pos.get("dte", ""),
                pos.get("entry_date", ""),
                pos.get("entry_premium", ""),
                pos.get("current_premium", ""),
                pnl,
                pos.get("unrealized_pnl_pct", ""),
                pos.get("take_profit_at", ""),
                pos.get("stop_loss_at", ""),
                pos.get("capital_required", ""),
                pos.get("status", ""),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)

                # Format currency columns
                if col_idx in (5, 9, 10, 11, 13, 14, 15):
                    cell.number_format = '$#,##0.00'
                # Format percentage columns
                elif col_idx == 12:
                    cell.number_format = '+0.0%;-0.0%;0.0%'

            _style_data_row(ws, row, len(POSITION_COLUMNS), pnl)

    wb.save(EXCEL_PATH)
    print(f"[ExcelLogger] Updated open positions in {EXCEL_PATH}")


# ---------------------------------------------------------------------------
# Log closed trade to Trade History
# ---------------------------------------------------------------------------

def log_closed_trade(closed_trade: dict):
    """
    Append a closed trade to the Trade History tab.
    Call this when a position is closed via paper_portfolio.close_position().
    """
    wb = init_workbook()
    ws = wb[HISTORY_TAB]

    pnl     = closed_trade.get("realized_pnl")
    pnl_pct = closed_trade.get("realized_pnl_pct", 0)
    win_loss = "WIN" if (pnl or 0) > 0 else "LOSS" if (pnl or 0) < 0 else "FLAT"

    row = ws.max_row + 1

    values = [
        closed_trade.get("position_id", ""),
        closed_trade.get("portfolio_name", "").title(),
        closed_trade.get("symbol", ""),
        closed_trade.get("option_type", ""),
        closed_trade.get("strategy", ""),
        closed_trade.get("strike", ""),
        closed_trade.get("expiration_date", ""),
        closed_trade.get("entry_date", ""),
        closed_trade.get("exit_date", ""),
        closed_trade.get("days_held", ""),
        closed_trade.get("entry_premium", ""),
        closed_trade.get("exit_premium", ""),
        pnl,
        pnl_pct,
        closed_trade.get("close_reason", ""),
        win_loss,
    ]

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        # Currency columns
        if col_idx in (6, 11, 12, 13):
            cell.number_format = '$#,##0.00'
        # Percentage column
        elif col_idx == 14:
            cell.number_format = '+0.0%;-0.0%;0.0%'

    _style_data_row(ws, row, len(HISTORY_COLUMNS), pnl)

    # Update title with running totals
    history = [r for r in ws.iter_rows(min_row=3, values_only=True) if r[0]]
    wins  = sum(1 for r in history if r[15] == "WIN")
    total = len(history)
    win_rate = f"{wins/total*100:.0f}%" if total else "0%"

    ws.cell(row=1, column=1).value = (
        f"Trade History — All Portfolios   "
        f"Total: {total}   Wins: {wins}   Win Rate: {win_rate}   "
        f"(Last updated: {_timestamp()})"
    )

    wb.save(EXCEL_PATH)
    print(f"[ExcelLogger] Logged closed trade {closed_trade.get('position_id')} to Trade History")


# ---------------------------------------------------------------------------
# Scan log (called by main.py after each scan)
# ---------------------------------------------------------------------------

def log_pipeline_run(symbol: str,
                     current_price: float,
                     eligible_puts: list,
                     eligible_calls: list,
                     portfolio_name: str = ""):
    """
    Append eligible options found this scan to the Scan Log tab.
    Signature matches existing main.py call — no changes needed there.
    Also triggers a refresh of open position tabs.
    """
    wb = init_workbook()
    ws = wb["Scan Log"]

    timestamp = _timestamp()

    rows_written = 0
    for put in eligible_puts:
        row = ws.max_row + 1
        values = [
            timestamp,
            portfolio_name or "—",
            symbol,
            "PUT",
            put.strike,
            put.bid,
            round(put.delta, 3) if put.delta else None,
            round(put.bid_ask_spread, 3),
            current_price,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx in (5, 6, 8, 9):
                cell.number_format = '$#,##0.00'
        _style_data_row(ws, row, len(SCAN_COLUMNS))
        rows_written += 1

    for call in eligible_calls:
        row = ws.max_row + 1
        values = [
            timestamp,
            portfolio_name or "—",
            symbol,
            "CALL",
            call.strike,
            call.bid,
            round(call.delta, 3) if call.delta else None,
            round(call.bid_ask_spread, 3),
            current_price,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx in (5, 6, 8, 9):
                cell.number_format = '$#,##0.00'
        _style_data_row(ws, row, len(SCAN_COLUMNS))
        rows_written += 1

    # Update scan log title with last run time
    ws.cell(row=1, column=1).value = (
        f"Scan Log — Eligible Options Found Each Run   "
        f"(Last scan: {timestamp})"
    )

    wb.save(EXCEL_PATH)

    if rows_written:
        print(f"[ExcelLogger] Logged {rows_written} rows to Scan Log")
    else:
        print("[ExcelLogger] No eligible options — nothing logged to Scan Log")

    # Refresh open positions tabs while we have the file open
    update_positions()
